import tkinter as tk
from tkinter import filedialog
import pandas as pd
import openpyxl

class PocketDCF:
    def __init__(self, root):
        self.root = root
        self.root.title("PocketDCF: University of Arizona Eller Finance")
        
        # Load the image
        self.background_image = tk.PhotoImage(file="./Images/EllerBackGround2.png")
        
        # Create a canvas with the same size as the image
        self.canvas = tk.Canvas(root, width=300, height=400)
        self.canvas.pack()
        
        # Add the image to the canvas
        self.canvas.create_image(0, 0, anchor=tk.NW, image=self.background_image)
        
        # Create labels and entry boxes for inputs
        labels = ["Revenue Growth", "COGS (% Rev)", "Op. Exp. (% Rev)", "Tax Rate", 
                  "Terminal Growth Rate", "Interest Rate"]
        self.inputs = {}
        for i, label in enumerate(labels):
            tk.Label(root, text=label).place(x=100, y=50+i*30)
            self.inputs[label] = tk.Entry(root, width=10)
            self.inputs[label].place(x=250, y=50+i*30)
        
        # Button to generate Excel sheets
        tk.Button(root, text="Generate Excel", command=self.generate_excel).place(x=400, y=350)
    
    def generate_excel(self):
        ticker = "TICKER"  # Placeholder for ticker symbol
        
        # Load the DCF template
        wb_template = openpyxl.load_workbook("./ExcelTemplates/DCF_TEMPLATE.xlsx")
        
        # Load the Income Statement, Balance Sheet, and Cash Flow files
        wb_income_statement = openpyxl.load_workbook("./FinancialStatements/CAVA_IS.xlsx")
        wb_balance_sheet = openpyxl.load_workbook("./FinancialStatements/CAVA_BS.xlsx")
        wb_cash_flow = openpyxl.load_workbook("./FinancialStatements/CAVA_CF.xlsx")
        
        # Get the sheets from the respective workbooks
        ws_income_statement = wb_income_statement.active
        ws_balance_sheet = wb_balance_sheet.active
        ws_cash_flow = wb_cash_flow.active
        
        # Get the sheets from the DCF template
        ws_template_income_statement = wb_template["Income Statement"]
        ws_template_balance_sheet = wb_template["Balance Sheet"]
        ws_template_cash_flow = wb_template["Statement Cash Flows"]
        
        # Merge data from Income Statement
        for row in ws_income_statement.iter_rows(min_row=2, values_only=True):
            ws_template_income_statement.append(row)
        
        # Merge data from Balance Sheet
        for row in ws_balance_sheet.iter_rows(min_row=2, values_only=True):
            ws_template_balance_sheet.append(row)
        
        # Merge data from Cash Flow
        for row in ws_cash_flow.iter_rows(min_row=2, values_only=True):
            ws_template_cash_flow.append(row)
        
        # Save the changes
        file_path = f"./GeneratedDCF/{ticker}_DCF.xlsx"
        wb_template.save(file_path)
        print("Data merged successfully! Saved as:", file_path)

if __name__ == "__main__":
    root = tk.Tk()
    app = PocketDCF(root)
    root.mainloop()
